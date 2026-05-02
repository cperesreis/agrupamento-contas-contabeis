"""
processamento.py
Módulo com toda a lógica de negócio:
  - ler_arquivo        : leitura, pré-filtro e validação
  - consolidar         : agrupamento e ordenação
  - classificar        : aplicação da base + classificações manuais
  - calcular_indicadores: totais e percentuais
  - exportar_ods       : geração do arquivo de saída
  - gerar_base_atualizada: merge da base com novas classificações
  - resetar_sessao     : limpeza de estado Streamlit
"""

import io
import os
import re
import pandas as pd

# Nome da base local usada automaticamente pelo app.py quando existir na pasta.
BASE_CLASSIFICACAO_NOME = "base_classificacao_atualizada.ods"
BASE_CLASSIFICACAO_ARQUIVO = os.path.join(os.path.dirname(os.path.abspath(__file__)), BASE_CLASSIFICACAO_NOME)

# Contas removidas ANTES de qualquer processamento
CONTAS_IGNORADAS_ORIGEM = {
    "DEVOLUCAO DE VENDAS",
    "MERCAD. EMITIDA P/ CONSERTO",
}

TIPOS_VALIDOS = {"C.OPERACIONAL", "NÃO OPERACIONAL", "IGNORAR"}


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _converter_valor(val) -> float:
    """Converte formatos monetários brasileiros e internacionais para float."""
    if pd.isna(val):
        raise ValueError(f"Valor nulo encontrado em VALPAGAMENTOTITULO.")
    s = str(val).strip()
    # Remove prefixo R$
    s = re.sub(r"R\$\s*", "", s)
    s = s.strip().replace(" ", "")

    # Formatos com separador de milhar e decimal: 1.234,56 ou 1,234.56
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    # Formato brasileiro sem milhar: 1234,56
    elif re.match(r"^\d+(,\d+)?$", s):
        s = s.replace(",", ".")
    # Formato americano com milhar: 1,234 ou 1,234.56
    elif re.match(r"^\d{1,3}(,\d{3})+(\.\d+)?$", s):
        s = s.replace(",", "")
    # Formato brasileiro só com milhar: 1.234
    elif re.match(r"^\d{1,3}(\.\d{3})+$", s):
        s = s.replace(".", "")
    # Formato 1234.56 já é válido
    try:
        return float(s)
    except ValueError:
        raise ValueError(f"Não foi possível converter o valor '{val}' para numérico.")


def _ler_ods(fileobj) -> pd.DataFrame:
    """Lê arquivo ODS usando pyexcel-ods3 ou odfpy via pandas."""
    try:
        import pyexcel_ods3
        data = pyexcel_ods3.get_data(fileobj)
        sheet_name = list(data.keys())[0]
        rows = data[sheet_name]
        if not rows:
            return pd.DataFrame()
        headers = [str(h) for h in rows[0]]
        return pd.DataFrame(rows[1:], columns=headers)
    except ImportError:
        pass

    # Fallback via odfpy + manual parse
    try:
        from odf.opendocument import load
        from odf.table import Table, TableRow, TableCell
        from odf.text import P

        fileobj.seek(0)
        doc = load(fileobj)
        tables = doc.spreadsheet.getElementsByType(Table)
        if not tables:
            return pd.DataFrame()
        rows_data = []
        for row in tables[0].getElementsByType(TableRow):
            row_vals = []
            for cell in row.getElementsByType(TableCell):
                ps = cell.getElementsByType(P)
                row_vals.append(str(ps[0]) if ps else "")
            rows_data.append(row_vals)
        if not rows_data:
            return pd.DataFrame()
        headers = [str(h) for h in rows_data[0]]
        return pd.DataFrame(rows_data[1:], columns=headers)
    except Exception as e:
        raise ValueError(f"Erro ao ler arquivo ODS: {e}")


def _ler_df(fileobj, nome_arquivo: str) -> pd.DataFrame:
    """Detecta extensão e lê o arquivo para DataFrame."""
    nome = nome_arquivo.lower()
    fileobj.seek(0)
    if nome.endswith(".csv"):
        try:
            return pd.read_csv(fileobj, sep=";", decimal=",", dtype=str)
        except Exception:
            fileobj.seek(0)
            return pd.read_csv(fileobj, dtype=str)
    elif nome.endswith(".xlsx"):
        return pd.read_excel(fileobj, dtype=str)
    elif nome.endswith(".xls"):
        return pd.read_excel(fileobj, dtype=str, engine="xlrd")
    elif nome.endswith(".ods"):
        return _ler_ods(fileobj)
    else:
        raise ValueError(f"Formato não suportado: {nome_arquivo}. Use ODS, XLSX, XLS ou CSV.")


# ─── Funções públicas ─────────────────────────────────────────────────────────

def ler_arquivo(file_upload, modo_base: bool = False):
    """
    Lê o arquivo de upload, aplica pré-filtro (modo_base=False) e valida colunas.

    Retorna:
        modo_base=False: (DataFrame, aviso_vazio: bool)
        modo_base=True : (DataFrame, lista_alertas: list)
    """
    file_upload.seek(0)
    df = _ler_df(file_upload, file_upload.name)

    # Normalizar nomes de colunas (strip de espaços)
    df.columns = [str(c).strip() for c in df.columns]

    if modo_base:
        alertas = []
        if "DESCRDEB" not in df.columns:
            raise ValueError("Base de classificação: coluna 'DESCRDEB' não encontrada.")
        if "TIPO DE CUSTO" not in df.columns:
            raise ValueError("Base de classificação: coluna 'TIPO DE CUSTO' não encontrada.")
        df = df[["DESCRDEB", "TIPO DE CUSTO"]].copy()
        df["DESCRDEB"] = df["DESCRDEB"].astype(str).str.strip()
        # Normalizar TIPO DE CUSTO: maiúsculas + strip
        df["TIPO DE CUSTO"] = df["TIPO DE CUSTO"].astype(str).str.strip().str.upper()
        # Detectar duplicatas com tipos diferentes
        dup = df[df.duplicated(subset=["DESCRDEB"], keep=False)]
        if not dup.empty:
            contas_dup = dup["DESCRDEB"].unique().tolist()
            for c in contas_dup:
                tipos = df[df["DESCRDEB"] == c]["TIPO DE CUSTO"].unique().tolist()
                if len(tipos) > 1:
                    alertas.append(
                        f"Conta duplicada na base com tipos diferentes: '{c}' → {tipos}. "
                        "Será tratada como PENDENTE DE CLASSIFICAÇÃO."
                    )
        return df, alertas

    # Planilha principal
    if "DESCRDEB" not in df.columns:
        raise ValueError("Coluna 'DESCRDEB' não encontrada na planilha.")
    if "VALPAGAMENTOTITULO" not in df.columns:
        raise ValueError("Coluna 'VALPAGAMENTOTITULO' não encontrada na planilha.")

    df = df[["DESCRDEB", "VALPAGAMENTOTITULO"]].copy()
    df["DESCRDEB"] = df["DESCRDEB"].astype(str).str.strip()

    # Pré-filtro: remover contas ignoradas desde a origem
    df = df[~df["DESCRDEB"].isin(CONTAS_IGNORADAS_ORIGEM)].copy()

    # Remover linhas com DESCRDEB vazio ou NaN
    df = df[df["DESCRDEB"].notna() & (df["DESCRDEB"] != "") & (df["DESCRDEB"] != "nan")]

    aviso_vazio = df.empty

    if not aviso_vazio:
        # Converter valores monetários
        valores_convertidos = []
        for idx, row in df.iterrows():
            try:
                valores_convertidos.append(_converter_valor(row["VALPAGAMENTOTITULO"]))
            except ValueError as e:
                raise ValueError(f"Linha {idx + 2}: {e}")
        df["VALPAGAMENTOTITULO"] = valores_convertidos

    return df, aviso_vazio


def ler_base_local(caminho: str = BASE_CLASSIFICACAO_ARQUIVO):
    """
    Lê a base de classificação salva localmente.

    Retorna o mesmo formato de ler_arquivo(..., modo_base=True):
        (DataFrame, lista_alertas)
    """
    with open(caminho, "rb") as arquivo:
        return ler_arquivo(arquivo, modo_base=True)


def consolidar(df: pd.DataFrame) -> pd.DataFrame:
    """Agrupa por DESCRDEB, soma valores e ordena decrescente."""
    if df.empty:
        return df.copy()
    df_cons = (
        df.groupby("DESCRDEB", as_index=False)["VALPAGAMENTOTITULO"]
        .sum()
        .sort_values("VALPAGAMENTOTITULO", ascending=False)
        .reset_index(drop=True)
    )
    return df_cons


def classificar(df: pd.DataFrame, df_base: pd.DataFrame | None, classificacoes_manuais: dict) -> tuple:
    """
    Aplica classificação usando base externa e classificações manuais.
    Retorna (df_classificado, alertas_duplicatas).
    """
    df = df.copy()
    alertas = []

    if df_base is not None and not df_base.empty:
        # Detectar duplicatas com tipos distintos e marcá-las como PENDENTE
        dup_contas = set()
        dup_check = df_base[df_base.duplicated(subset=["DESCRDEB"], keep=False)]
        for conta in dup_check["DESCRDEB"].unique():
            tipos = df_base[df_base["DESCRDEB"] == conta]["TIPO DE CUSTO"].unique()
            if len(tipos) > 1:
                dup_contas.add(conta)

        # Montar mapeamento sem duplicatas conflitantes
        df_base_valida = df_base[~df_base["DESCRDEB"].isin(dup_contas)].copy()
        # Última ocorrência para não-duplicatas (seguro)
        mapa_base = df_base_valida.drop_duplicates(subset=["DESCRDEB"], keep="last").set_index("DESCRDEB")["TIPO DE CUSTO"].to_dict()
    else:
        mapa_base = {}
        dup_contas = set()

    def classificar_conta(conta: str) -> str:
        # 1. Classificação manual tem precedência
        if conta in classificacoes_manuais:
            return classificacoes_manuais[conta]
        # 2. Conta duplicada na base → PENDENTE
        if conta in dup_contas:
            return "PENDENTE DE CLASSIFICAÇÃO"
        # 3. Base externa
        if conta in mapa_base:
            tipo = mapa_base[conta]
            return tipo if tipo in TIPOS_VALIDOS else "PENDENTE DE CLASSIFICAÇÃO"
        # 4. Não encontrada
        return "PENDENTE DE CLASSIFICAÇÃO"

    df["TIPO DE CUSTO"] = df["DESCRDEB"].apply(classificar_conta)

    # Calcular % sobre faturamento (será preenchido em calcular_indicadores)
    df["% SOBRE FATURAMENTO"] = None

    return df, alertas


def calcular_indicadores(df: pd.DataFrame, faturamento: float) -> dict:
    """Calcula totais e percentuais. Também preenche a coluna % SOBRE FATURAMENTO."""
    if faturamento <= 0:
        raise ValueError("Faturamento deve ser maior que zero.")

    total_despesas = df["VALPAGAMENTOTITULO"].sum()
    custo_op = df.loc[df["TIPO DE CUSTO"] == "C.OPERACIONAL", "VALPAGAMENTOTITULO"].sum()
    saldo = faturamento - total_despesas

    # Preencher percentual por conta
    df["% SOBRE FATURAMENTO"] = (df["VALPAGAMENTOTITULO"] / faturamento) * 100

    return {
        "faturamento": faturamento,
        "total_despesas": total_despesas,
        "custo_operacional": custo_op,
        "saldo": saldo,
        "pct_despesas": (total_despesas / faturamento) * 100,
        "pct_operacional": (custo_op / faturamento) * 100,
        "pct_saldo": (saldo / faturamento) * 100,
    }


def exportar_ods(df: pd.DataFrame, indicadores: dict, faturamento: float) -> tuple:
    """
    Gera o arquivo ODS consolidado e (opcionalmente) o de pendências.
    Retorna (ods_bytes, ods_pendencias_bytes | None).
    """
    try:
        from odf.opendocument import OpenDocumentSpreadsheet
        from odf.table import Table, TableRow, TableCell
        from odf.text import P
        from odf.style import Style, TextProperties, TableCellProperties
        import odf.number as odfnum

        def _make_doc():
            return OpenDocumentSpreadsheet()

        def _add_number_styles(doc):
            moeda = odfnum.CurrencyStyle(name="moedaBRL")
            moeda.addElement(odfnum.CurrencySymbol(text="R$"))
            moeda.addElement(odfnum.Text(text=" "))
            moeda.addElement(odfnum.Number(decimalplaces=2, minintegerdigits=1, grouping=True))
            doc.automaticstyles.addElement(moeda)

            percentual = odfnum.PercentageStyle(name="percentual")
            percentual.addElement(odfnum.Number(decimalplaces=2, minintegerdigits=1))
            percentual.addElement(odfnum.Text(text="%"))
            doc.automaticstyles.addElement(percentual)

        def _add_style(
            doc,
            name,
            bold=False,
            italic=False,
            font_size=None,
            color=None,
            bg_color=None,
            data_style_name=None,
        ):
            attrs = {"name": name, "family": "table-cell"}
            if data_style_name:
                attrs["datastylename"] = data_style_name
            style = Style(**attrs)
            text_props = {}
            if bold:
                text_props["fontweight"] = "bold"
            if italic:
                text_props["fontstyle"] = "italic"
            if font_size:
                text_props["fontsize"] = font_size
            if color:
                text_props["color"] = color
            if text_props:
                style.addElement(TextProperties(**text_props))
            if bg_color:
                style.addElement(TableCellProperties(backgroundcolor=bg_color))
            doc.automaticstyles.addElement(style)
            return name

        def _cell_text(value, style_name=None):
            tc = TableCell()
            if style_name:
                tc.setAttribute("stylename", style_name)
            tc.setAttribute("valuetype", "string")
            p = P(text=str(value) if value is not None else "")
            tc.addElement(p)
            return tc

        def _cell_float(value, style_name=None, display_value=None):
            tc = TableCell()
            if style_name:
                tc.setAttribute("stylename", style_name)
            if value is None or pd.isna(value):
                return tc
            numero = float(value)
            tc.setAttribute("valuetype", "float")
            tc.setAttribute("value", str(numero))
            tc.addElement(P(text=str(display_value if display_value is not None else numero)))
            return tc

        def _pct_decimal(value):
            if value is None or pd.isna(value):
                return None
            return float(value) / 100

        def _empty_row(table, n_cols):
            tr = TableRow()
            for _ in range(n_cols):
                tr.addElement(TableCell())
            table.addElement(tr)

        # ── Documento principal ───────────────────────────────────────────────
        doc = _make_doc()
        _add_number_styles(doc)

        s_header  = _add_style(doc, "sHeader",  bold=True)
        s_normal  = _add_style(doc, "sNormal")
        s_ignorar = _add_style(doc, "sIgnorar", italic=True, color="#888888")
        s_total   = _add_style(doc, "sTotal",   bold=True, font_size="16pt")
        s_moeda = _add_style(doc, "sMoeda", data_style_name="moedaBRL")
        s_pct = _add_style(doc, "sPercentual", data_style_name="percentual")
        s_moeda_ignorar = _add_style(doc, "sMoedaIgnorar", italic=True, color="#888888", data_style_name="moedaBRL")
        s_pct_ignorar = _add_style(doc, "sPercentualIgnorar", italic=True, color="#888888", data_style_name="percentual")
        s_total_moeda = _add_style(doc, "sTotalMoeda", bold=True, font_size="16pt", data_style_name="moedaBRL")
        s_total_pct = _add_style(doc, "sTotalPercentual", bold=True, font_size="16pt", data_style_name="percentual")
        s_total_vermelho = _add_style(doc, "sTotalVermelho", bold=True, font_size="16pt", color="#dc2626")
        s_total_moeda_vermelho = _add_style(
            doc,
            "sTotalMoedaVermelho",
            bold=True,
            font_size="16pt",
            color="#dc2626",
            data_style_name="moedaBRL",
        )
        s_total_pct_vermelho = _add_style(
            doc,
            "sTotalPercentualVermelho",
            bold=True,
            font_size="16pt",
            color="#dc2626",
            data_style_name="percentual",
        )

        table = Table(name="CONSOLIDADO")
        doc.spreadsheet.addElement(table)

        # Cabeçalho
        tr = TableRow()
        for col in ["DESCRDEB", "VALPAGAMENTOTITULO", "TIPO DE CUSTO", "% SOBRE FATURAMENTO"]:
            tr.addElement(_cell_text(col, s_header))
        table.addElement(tr)

        # Dados
        for _, row in df.iterrows():
            tipo = row["TIPO DE CUSTO"]
            estilo = s_ignorar if tipo == "IGNORAR" else s_normal
            estilo_moeda = s_moeda_ignorar if tipo == "IGNORAR" else s_moeda
            estilo_pct = s_pct_ignorar if tipo == "IGNORAR" else s_pct
            tr = TableRow()
            tr.addElement(_cell_text(row["DESCRDEB"], estilo))
            tr.addElement(_cell_float(row["VALPAGAMENTOTITULO"], estilo_moeda))
            tr.addElement(_cell_text(tipo, estilo))
            tr.addElement(_cell_float(_pct_decimal(row.get("% SOBRE FATURAMENTO")), estilo_pct))
            table.addElement(tr)

        # 4 linhas em branco
        for _ in range(4):
            _empty_row(table, 4)

        # Bloco de totais
        ind = indicadores
        totais = [
            ("FATURAMENTO DO PERÍODO",      ind["faturamento"],      ind["pct_saldo"] * 0),
            ("TOTAL GERAL DE DESPESAS",     ind["total_despesas"],   ind["pct_despesas"]),
            ("SALDO",                        ind["saldo"],            ind["pct_saldo"]),
            ("TOTAL DE CUSTO OPERACIONAL",  ind["custo_operacional"],ind["pct_operacional"]),
        ]
        for label, valor, pct in totais:
            despesa_total = label == "TOTAL GERAL DE DESPESAS"
            saldo_negativo = label == "SALDO" and valor < 0
            usar_vermelho = despesa_total or saldo_negativo
            estilo_total = s_total_vermelho if usar_vermelho else s_total
            estilo_moeda_total = s_total_moeda_vermelho if usar_vermelho else s_total_moeda
            estilo_pct_total = s_total_pct_vermelho if saldo_negativo else s_total_pct
            valor_exibicao = -abs(valor) if despesa_total else valor
            tr = TableRow()
            tr.addElement(_cell_text(label, estilo_total))
            tr.addElement(_cell_float(valor_exibicao, estilo_moeda_total))
            tr.addElement(_cell_text("", estilo_total))
            tr.addElement(_cell_float(_pct_decimal(pct), estilo_pct_total) if label != "FATURAMENTO DO PERÍODO" else _cell_text("", estilo_total))
            table.addElement(tr)

        buf = io.BytesIO()
        doc.save(buf)
        ods_bytes = buf.getvalue()

        # ── Aba / arquivo de pendências ───────────────────────────────────────
        df_pend = df[df["TIPO DE CUSTO"] == "PENDENTE DE CLASSIFICAÇÃO"].copy()
        ods_pend_bytes = None

        if not df_pend.empty:
            doc_p = _make_doc()
            _add_number_styles(doc_p)
            sp = _add_style(doc_p, "sHeader", bold=True)
            sn = _add_style(doc_p, "sNormal")
            sm = _add_style(doc_p, "sMoeda", data_style_name="moedaBRL")

            tp = Table(name="PENDENTES")
            doc_p.spreadsheet.addElement(tp)

            tr = TableRow()
            tr.addElement(_cell_text("DESCRDEB", sp))
            tr.addElement(_cell_text("VALOR", sp))
            tp.addElement(tr)

            for _, row in df_pend.sort_values("VALPAGAMENTOTITULO", ascending=False).iterrows():
                tr = TableRow()
                tr.addElement(_cell_text(row["DESCRDEB"], sn))
                tr.addElement(_cell_float(row["VALPAGAMENTOTITULO"], sm))
                tp.addElement(tr)

            buf_p = io.BytesIO()
            doc_p.save(buf_p)
            ods_pend_bytes = buf_p.getvalue()

        return ods_bytes, ods_pend_bytes

    except ImportError:
        # Fallback: usar openpyxl para gerar XLSX caso odfpy não esteja disponível
        return _exportar_xlsx_fallback(df, indicadores)


def _exportar_xlsx_fallback(df, indicadores):
    """Fallback: gera XLSX quando odfpy não está disponível."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "CONSOLIDADO"

    header_font = Font(bold=True)
    total_font  = Font(bold=True, size=18)
    total_red_font = Font(bold=True, size=18, color="DC2626")
    gray_font   = Font(color="888888", italic=True)

    headers = ["DESCRDEB", "VALPAGAMENTOTITULO", "TIPO DE CUSTO", "% SOBRE FATURAMENTO"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font

    for _, row in df.iterrows():
        pct = row.get("% SOBRE FATURAMENTO")
        ws.append([
            row["DESCRDEB"],
            float(row["VALPAGAMENTOTITULO"]),
            row["TIPO DE CUSTO"],
            float(pct) / 100 if pd.notna(pct) else None,
        ])
        ws.cell(ws.max_row, 2).number_format = 'R$ #,##0.00'
        ws.cell(ws.max_row, 4).number_format = '0.00%'
        if row["TIPO DE CUSTO"] == "IGNORAR":
            for cell in ws[ws.max_row]:
                cell.font = gray_font

    for _ in range(4):
        ws.append(["", "", "", ""])

    ind = indicadores
    for label, valor, pct in [
        ("FATURAMENTO DO PERÍODO",     ind["faturamento"],      None),
        ("TOTAL GERAL DE DESPESAS",    ind["total_despesas"],   ind["pct_despesas"] / 100),
        ("SALDO",                       ind["saldo"],            ind["pct_saldo"] / 100),
        ("TOTAL DE CUSTO OPERACIONAL", ind["custo_operacional"],ind["pct_operacional"] / 100),
    ]:
        despesa_total = label == "TOTAL GERAL DE DESPESAS"
        saldo_negativo = label == "SALDO" and valor < 0
        usar_vermelho = despesa_total or saldo_negativo
        ws.append([label, -abs(float(valor)) if despesa_total else float(valor), "", pct])
        ws.cell(ws.max_row, 2).number_format = 'R$ #,##0.00'
        ws.cell(ws.max_row, 4).number_format = '0.00%'
        for cell in ws[ws.max_row]:
            cell.font = total_red_font if usar_vermelho else total_font
        if despesa_total:
            ws.cell(ws.max_row, 4).font = total_font

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), None


def gerar_base_atualizada(df_base_original, classificacoes_manuais: dict) -> bytes:
    """
    Mescla a base original com novas classificações manuais, sem duplicatas.
    Retorna bytes ODS.
    """
    rows_novos = [
        {"DESCRDEB": k, "TIPO DE CUSTO": v}
        for k, v in classificacoes_manuais.items()
    ]
    df_novos = pd.DataFrame(rows_novos)

    if df_base_original is not None and not df_base_original.empty:
        df_merged = pd.concat([df_base_original, df_novos], ignore_index=True)
        # Em caso de duplicata, a classificação manual (última) prevalece
        df_merged = df_merged.drop_duplicates(subset=["DESCRDEB"], keep="last")
    else:
        df_merged = df_novos

    df_merged = df_merged.sort_values("DESCRDEB").reset_index(drop=True)

    try:
        from odf.opendocument import OpenDocumentSpreadsheet
        from odf.table import Table, TableRow, TableCell
        from odf.text import P
        from odf.style import Style, TextProperties

        doc = OpenDocumentSpreadsheet()
        s_h = Style(name="h", family="table-cell")
        s_h.addElement(TextProperties(fontweight="bold"))
        doc.automaticstyles.addElement(s_h)

        def _cell(val, style=None):
            tc = TableCell(valuetype="string")
            if style:
                tc.setAttribute("stylename", style)
            tc.addElement(P(text=str(val)))
            return tc

        tbl = Table(name="BASE_CLASSIFICACAO")
        doc.spreadsheet.addElement(tbl)

        tr = TableRow()
        tr.addElement(_cell("DESCRDEB",     "h"))
        tr.addElement(_cell("TIPO DE CUSTO","h"))
        tbl.addElement(tr)

        for _, row in df_merged.iterrows():
            tr = TableRow()
            tr.addElement(_cell(row["DESCRDEB"]))
            tr.addElement(_cell(row["TIPO DE CUSTO"]))
            tbl.addElement(tr)

        buf = io.BytesIO()
        doc.save(buf)
        return buf.getvalue()

    except ImportError:
        # Fallback XLSX
        from openpyxl import Workbook
        from openpyxl.styles import Font
        wb = Workbook()
        ws = wb.active
        ws.title = "BASE_CLASSIFICACAO"
        ws.append(["DESCRDEB", "TIPO DE CUSTO"])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for _, row in df_merged.iterrows():
            ws.append([row["DESCRDEB"], row["TIPO DE CUSTO"]])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()


def resetar_sessao(session_state) -> None:
    """Limpa variáveis de estado ao detectar novo arquivo de upload."""
    session_state.classificacoes_manuais = {}
    session_state.df_consolidado = None
    session_state.indicadores = None
    session_state.processado = False
    session_state.ods_bytes = None
    session_state.ods_pendencias_bytes = None
    session_state.ods_base_bytes = None
    session_state.pop("perguntar_salvar_base", None)
    session_state.pop("df_base_original", None)
    session_state.pop("faturamento", None)
